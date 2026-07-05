import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('Template Jadual OT.xlsx')
ws = wb['Formatted_Readonly_Kemaskini_20']

lines = []

def log(s):
    lines.append(s)

log('=== ROW 1 (Title) ===')
c = ws.cell(1,1)
fg = c.fill.fgColor.rgb if c.fill.fgColor else 'none'
fc = c.font.color.rgb if c.font.color else 'none'
log(f'Font: name={c.font.name}, sz={c.font.size}, bold={c.font.bold}, color={fc}')
log(f'Fill: type={c.fill.fill_type}, fg={fg}')
log(f'Align: h={c.alignment.horizontal}, v={c.alignment.vertical}, wrap={c.alignment.wrap_text}')

log('\n=== ROW 2 (Cat headers) ===')
for col in [1,2,3,7,19]:
    c = ws.cell(2, col)
    fg = c.fill.fgColor.rgb if c.fill.fgColor else 'none'
    fc = c.font.color.rgb if c.font.color else 'none'
    log(f'Col{col}: val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}, sz={c.font.size}')

log('\n=== ROW 3 (Times) ===')
for col in [1,2,3,7,10,12,14,16,19]:
    c = ws.cell(3, col)
    fg = c.fill.fgColor.rgb if c.fill.fgColor else 'none'
    fc = c.font.color.rgb if c.font.color else 'none'
    log(f'Col{col}: val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}')

log('\n=== ROW 4 (Sub-headers) ===')
for col in [1,2,3,6,7,11,12,15,16,19,21]:
    c = ws.cell(4, col)
    fg = c.fill.fgColor.rgb if c.fill.fgColor else 'none'
    fc = c.font.color.rgb if c.font.color else 'none'
    log(f'Col{col}: val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}')

log('\n=== ROW 5 (Slot names) ===')
for col in range(1, 23):
    c = ws.cell(5, col)
    fg = c.fill.fgColor.rgb if c.fill.fgColor else 'none'
    fc = c.font.color.rgb if c.font.color else 'none'
    log(f'Col{col}: val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}, sz={c.font.size}')

log('\n=== DATA ROWS (6=Holiday, 9=Weekday) ===')
for r in [6, 9, 11]:
    log(f'--- Row {r} ---')
    for col in [1,2,3,4,5,6,7,8,9,10,11,20,22]:
        c = ws.cell(r, col)
        fg = c.fill.fgColor.rgb if c.fill.fgColor else 'none'
        fc = c.font.color.rgb if c.font.color else 'none'
        if c.value or col <= 2:
            log(f'  Col{col}: val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}')

log('\n=== COLUMN WIDTHS ===')
for i in range(1, 23):
    letter = get_column_letter(i)
    w = ws.column_dimensions[letter].width
    log(f'{letter}: {w}')

log('\n=== ROW HEIGHTS ===')
for i in range(1, 8):
    log(f'Row {i}: {ws.row_dimensions[i].height}')

log('\n=== MERGED CELLS ===')
for m in ws.merged_cells.ranges:
    log(str(m))

log('\n=== BORDER SAMPLES ===')
for r, c_idx in [(1,1), (2,1), (5,1), (6,1), (6,3), (9,3)]:
    c = ws.cell(r, c_idx)
    b = c.border
    ts = b.top.style if b.top else 'none'
    bs = b.bottom.style if b.bottom else 'none'
    ls = b.left.style if b.left else 'none'
    rs = b.right.style if b.right else 'none'
    log(f'R{r}C{c_idx}: top={ts}, bottom={bs}, left={ls}, right={rs}')

with open('template_analysis.txt', 'w') as f:
    f.write('\n'.join(lines))

print('Analysis written to template_analysis.txt')