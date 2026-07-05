import openpyxl
wb = openpyxl.load_workbook('Template Jadual OT.xlsx')
ws = wb['Formatted_Readonly_Kemaskini_20']
lines = []
def log(s): lines.append(s)

# Check actual cell values for row 2 (category headers)
log('=== ROW 2 FULL SCAN ===')
for col in range(1, 24):
    c = ws.cell(2, col)
    if c.value:
        fg = c.fill.fgColor.rgb if c.fill.fgColor and hasattr(c.fill.fgColor, 'rgb') else str(c.fill.fgColor)
        log(f'  Col{col}({openpyxl.utils.get_column_letter(col)}): val={c.value!r}, fill_fg={fg}')

# Check row 3 (time headers) with actual fill colors
log('\n=== ROW 3 FULL SCAN ===')
for col in range(1, 24):
    c = ws.cell(3, col)
    if c.value:
        fg = c.fill.fgColor.rgb if c.fill.fgColor and hasattr(c.fill.fgColor, 'rgb') else str(c.fill.fgColor)
        fc = c.font.color.rgb if c.font.color and hasattr(c.font.color, 'rgb') else str(c.font.color)
        log(f'  Col{col}({openpyxl.utils.get_column_letter(col)}): val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}, sz={c.font.size}')

# Check row 4 (sub-headers) with fills
log('\n=== ROW 4 FULL SCAN ===')
for col in range(1, 24):
    c = ws.cell(4, col)
    if c.value:
        fg = c.fill.fgColor.rgb if c.fill.fgColor and hasattr(c.fill.fgColor, 'rgb') else str(c.fill.fgColor)
        fc = c.font.color.rgb if c.font.color and hasattr(c.font.color, 'rgb') else str(c.font.color)
        log(f'  Col{col}({openpyxl.utils.get_column_letter(col)}): val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}')

# Data row formatting - check ALL cells in a data row
log('\n=== ROW 6 (Holiday data) ALL COLS ===')
for col in range(1, 24):
    c = ws.cell(6, col)
    fg = c.fill.fgColor.rgb if c.fill.fgColor and hasattr(c.fill.fgColor, 'rgb') else str(c.fill.fgColor)
    fc = c.font.color.rgb if c.font.color and hasattr(c.font.color, 'rgb') else str(c.font.color)
    b = c.border
    bs = b.bottom.style if b.bottom else 'none'
    log(f'  Col{col}({openpyxl.utils.get_column_letter(col)}): val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}, border_bottom={bs}')

log('\n=== ROW 9 (Weekday data) ALL COLS ===')
for col in range(1, 24):
    c = ws.cell(9, col)
    fg = c.fill.fgColor.rgb if c.fill.fgColor and hasattr(c.fill.fgColor, 'rgb') else str(c.fill.fgColor)
    fc = c.font.color.rgb if c.font.color and hasattr(c.font.color, 'rgb') else str(c.font.color)
    b = c.border
    bs = b.bottom.style if b.bottom else 'none'
    log(f'  Col{col}({openpyxl.utils.get_column_letter(col)}): val={c.value!r}, fill_fg={fg}, font_color={fc}, bold={c.font.bold}, border_bottom={bs}')

# Footer rows
log('\n=== FOOTER ROWS ===')
for r in range(37, 42):
    for col in [2, 20]:
        c = ws.cell(r, col)
        if c.value:
            log(f'  R{r}C{col}: val={c.value!r}, bold={c.font.bold}, italic={c.font.italic}, sz={c.font.size}')

with open('template_analysis2.txt', 'w') as f:
    f.write('\n'.join(lines))
print('Done')